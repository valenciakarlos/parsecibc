#!/usr/bin/python
# References:
# https://github.com/python-excel/tutorial/blob/master/python-excel.pdf
# https://blogs.harvard.edu/rprasad/2014/06/16/reading-excel-with-python-xlrd/
# Using a new library as xlwt does not support SUMIFS
# Reference: https://openpyxl.readthedocs.io/en/default/usage.html
from xlrd import open_workbook
import sys
from datetime import datetime # Library to manage times
# Reference for datetime : https://docs.python.org/2/library/datetime.html#datetime-objects
# for directory manipulation

import os.path
from xlrd.sheet import ctype_text  

def dollar_format(cell):
    cell.number_format='_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
    
def pct_format(cell):
    cell.number_format="0.0%"

def show_column_names(xl_sheet):
    row = xl_sheet.row(0)  # 1st row
    print(60*'-' + 'n(Column #) value [type]n' + 60*'-')
    for idx, cell_obj in enumerate(row):
        cell_type_str = ctype_text.get(cell_obj.ctype, 'unknown type')
        print('(%s) %s [%s]' % (idx, cell_obj.value, cell_type_str, ))

# Method to get opening price off Google finance. Temporary hack as API doesnt seem to be working. Code exception handling if possible
def getOpQuote(symbol):
 import json
 import requests
 req="https://finance.google.com/finance?q="+symbol+"&output=json"
 rsp = requests.get(req)
 if rsp.status_code in (200,):
    fin_data = json.loads(rsp.content[6:-2].decode('unicode_escape'))
    # Want closing price but dont know where. I think I can parse without the json output
    return fin_data['op']

# Method to write headers to a sheet
def add_headers(ws,google):
 ws['A1']="Stock"					# COL=1
 ws['B1']="Qty"						# COL=2
 ws['C1']="Dividends"			# COL=3
 ws['D1']="Invested"				# COL=4
 ws['E1']="Comissions"			# COL=5
 ws['F1']="Taxes"					# COL=6
 ws['G1']="DivReturn"			# COL=7
 ws['H1']="FQ_Stock"					# COL=8
 ws['I1']="Current Price"	# COL=9
 ws['J1']="Current Price (CAD)"  # COL=10
 ws['K1']="Current Amount (CAD)" # COL=11
 ws['L1']="Gain if sold "  			# COL=12
 ws['M1']="Return if sold "  		# COL=13
 if google: # Flag to indicate if the sheet if for google finance
  ws['N1']="Full Name"

#
# Adjust column width automatically
def adjust_col(col):
    max_length=0
    column=col[0].column # Gets the column name
    for cell in col:     # Traverse the cells and see their length
        try:
	   if len(str(cell.value)) > max_length:
	      max_length = len(cell.value)
	except:
	   pass
    #print "Max length for column "+column +" is :"+str(max_length)
    adjusted_width=(max_length + 2) * 1.2
    return adjusted_width
    #print "Sugested width for colum "+column +" is :"+str(adjusted_width)

# Method to adjust column width
def adjust_headers(ws):
 from openpyxl.styles import Alignment
 for col in ws.columns:
  column = col[0].column # Extracts the column name
  header_cell=col[0] # Extracts the header only 
  adjusted_width=adjust_col(col)
  if (adjusted_width<=10):
     ws.column_dimensions[column].width=adjusted_width
  elif (adjusted_width>=20):
     ws.column_dimensions[column].width=14 # Sets width for the column. Hack but let's see
  # Here we should get the column and decide accordingly. Maybe have it on a dictionary already store.
  header_cell.alignment=Alignment(horizontal='center',wrapText=True) # Sets attributes for the header cell
  
# Method to write the totals
def write_totals(ws):
  num_rows=ws.max_row
  totals_row=ws.max_row+2
  #print "Will write totals on row # "+str(totals_row)

  # COL=1 TOTALS title
  ws.cell(row=totals_row,column=1,value="TOTALS")

  # COL=3 dividends colums. "C"
  sum_str="=SUM(C2:C"+str(num_rows)+")"
  ws.cell(row=totals_row,column=3,value=sum_str)
  dollar_format(ws.cell(row=totals_row,column=3))

  # COL=4 Total Invested. "D"
  sum_str="=SUM(D2:D"+str(num_rows)+")"
  ws.cell(row=totals_row,column=4,value=sum_str)
  dollar_format(ws.cell(row=totals_row,column=4))

  # COL=5 Comisions paid. "E"
  sum_str="=SUM(E2:E"+str(num_rows)+")"
  ws.cell(row=totals_row,column=5,value=sum_str)
  dollar_format(ws.cell(row=totals_row,column=5))

  # COL=6 Taxes paid. "F"
  sum_str="=SUM(F2:F"+str(num_rows)+")"
  ws.cell(row=totals_row,column=6,value=sum_str)
  dollar_format(ws.cell(row=totals_row,column=6))

  # COL=7 Dividends return. "G" Calculated based on the dividends vs. invested
  div_return_str="=C"+str(totals_row)+"/(-D"+str(totals_row)+"-F"+str(totals_row)+"+E"+str(totals_row)+")"
  ws.cell(row=totals_row,column=7,value=div_return_str )
  pct_format(ws.cell(row=totals_row,column=7))

  # COL=11 Current amount "K"
  sum_str="=SUM(K2:K"+str(num_rows)+")"
  ws.cell(row=totals_row,column=11,value=sum_str)
  dollar_format(ws.cell(row=totals_row,column=11))

  # COL=12. Total gains if sold "L"
  sum_str="=SUM(L2:L"+str(num_rows)+")"
  ws.cell(row=totals_row,column=12,value=sum_str)
  dollar_format(ws.cell(row=totals_row,column=12))

  # COL=13. Total return if sold "M"
  return_str="=L"+str(totals_row)+"/(-D"+str(totals_row)+"+E"+str(totals_row)+")"
  ws.cell(row=totals_row,column=13,value=return_str)
  pct_format(ws.cell(row=totals_row,column=13))





# Main

if (len(sys.argv) == 2):
  filename=sys.argv[1]
  print "Reading file " + os.path.basename(filename) + " Located on path " + os.path.dirname(filename)
  book=open_workbook(filename)
  
else:
  print "Use %s <filename>" %(sys.argv[0])
  sys.exit()

sheet=book.sheet_by_index(0)
print "Name =" + sheet.name
print "Rows=" + str(sheet.nrows)
print "Cols=" + str(sheet.ncols)

############### CONSTANTS FOR PARSING ###############################
COL_TRANS_TYPE="D"
COL_SYMBOL="E"
COL_MARKET="F"
COL_QTY="H"
COL_PRICE="J"
COL_COM="K"
COL_AMOUNT="N"
#####################################################################
#### Comision constant ##############
COMISION="-7"
############# VARIABLES FOR PARSING #################################
num_cols = sheet.ncols   # Number of columns
num_rows = sheet.nrows
#####################################################################
# Load up the whole spreadsheet from row #9 (where transactions with header starts)
data = [sheet.row_values(i) for i in xrange(9,sheet.nrows)]
labels = data[0]    # Don't sort our headers
data = data[1:]     # Data begins on the second row
# Order the data based on the 1st column
#data.sort(key=lambda x: x[0])
data.sort(key=lambda x: x[0],reverse=True)
#
#
# Print 1st column
flag=0
for row_idx in range(0, sheet.nrows):    # Iterate through rows
    #print ('-'*40)
    #print ('Row: %s' % row_idx)   # Print row number
    cell_obj = sheet.cell(row_idx, 0)  
    if (flag):
       tx_date=sheet.cell(row_idx, 0).value
       tx_settlement=sheet.cell(row_idx, 1).value
       tx_type=sheet.cell(row_idx,3).value
       symbol=sheet.cell(row_idx,4).value
       market=sheet.cell(row_idx,5).value
       qty=sheet.cell(row_idx,7).value
       cur_price=sheet.cell(row_idx,8).value
       price=sheet.cell(row_idx,9).value
       comission=sheet.cell(row_idx,10).value
       exc_rate=sheet.cell(row_idx,11).value
       cur_amount=sheet.cell(row_idx,12).value
       amount=sheet.cell(row_idx,13).value

       # Convert to date object
       dtobj=datetime.strptime(tx_date, "%B %d, %Y")  # https://docs.python.org/2/library/datetime.html#strftime-strptime-behavior

       #print ('[%s] [%s] [%s] [%s]' % (dtobj.strftime("%b %d,%y"),tx_type,symbol,amount))
       #print ('[%s] [%s] [%s] [%s] [%s] [%s] [%s] [%s] [%s] [%s] [%s] [%s]' % (tx_date,tx_settlement,tx_type,symbol,market,qty,cur_price,price,comission, exc_rate,cur_amount,amount))

    if cell_obj.value=="Transaction Date":
       print "Title found on row Printing details of transactions %s" %(row_idx+1)
       flag=1

# Print results
import xlwt
# Trying to print the date with format
# 
from xlwt import *
#from xlwt import XFStyle
style = XFStyle()
#style.num_format_str='%B-%d-%Y'
#style.num_format_str='DD-MM-YY'
style.num_format_str='D-MMM-YY'
# Check formats at : https://github.com/python-excel/xlwt/blob/master/examples/num_formats.py


bk = xlwt.Workbook()
sheet = bk.add_sheet(sheet.name)

for idx, label in enumerate(labels):
     sheet.write(0, idx, label)

for idx_r, row in enumerate(data):
    for idx_c, value in enumerate(row):
        if idx_c == 0 or idx_c==1 :
           dtobj=datetime.strptime(value, "%B %d, %Y")
           sheet.write(idx_r+1, idx_c, dtobj, style)
        else:
           sheet.write(idx_r+1, idx_c, value)

bk.save('/Users/valenc6/Documents/Investments/result.xls')

bk=open_workbook("/Users/valenc6/Documents/Investments/result.xls")
sheet=bk.sheet_by_index(0)
# Order the spread sheet
# Load up the whole spreadsheet from row #0 (where transactions with header starts)
data = [sheet.row_values(i) for i in xrange(0,sheet.nrows)]
labels = data[0]    # Don't sort our headers
data = data[1:]     # Data begins on the second row
# Order the data based on the 1st column
data.sort(key=lambda x: x[0],reverse=False)

# Extract a list of the symbols 4 is the column with the symbols
sublist=[sheet.cell_value(row,4) for row in range(1,sheet.nrows)]
# Extract a list of the market corresponding to each symbol. 5 is the column
marketlist=[sheet.cell_value(row,5) for row in range(1,sheet.nrows)]
# sublist is now an array with all the symbols and marketlist has its corresponding market

symbols_list=[]
dict_symbols={}
for idx,value in enumerate(sublist):
  if (value not in symbols_list) and (value) :
     symbols_list.append(value)
     if marketlist[idx]=="CDN":
        dict_symbols[value]="TSE"
     elif marketlist[idx]=="US":
        dict_symbols[value]="NYSEARCA"
     else:
        dict_symbols[value]=marketlist[idx]

print "Dictionary of symbols is now:"
print dict_symbols

# Results are now on a list called symbols_list
###########################################################################
# Now determine the ranges to use for the Calculations work sheet
############# VARIABLES FOR PARSING #################################
num_cols = sheet.ncols   # Number of columns
num_rows = sheet.nrows
#####################################################################
#
#
symbols_range="Ordered!$"+COL_SYMBOL+"$2:$"+COL_SYMBOL+"$"+str(num_rows)
qty_range="Ordered!$"+COL_QTY+"$2:$"+COL_QTY+"$"+str(num_rows)
amount_range="Ordered!$"+COL_AMOUNT+"$2:$"+COL_AMOUNT+"$"+str(num_rows)
trans_type_range="Ordered!$"+COL_TRANS_TYPE+"$2:$"+COL_TRANS_TYPE+"$"+str(num_rows) 
com_range="Ordered!$"+COL_COM+"$2:$"+COL_COM+"$"+str(num_rows) 
###########################################################################
#Write new spreadsheet 

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

bk = Workbook()
# Note here sheet changed type. It is now a openpyxl  sheet. 
sheet=bk.active
sheet.title="Ordered"

# write column titles
for idx, label in enumerate(labels):
     #print "Writting to col=%s row=%s val=%s" %(str(idx+1),str(1),str(label))
     sheet.cell(column=idx+1, row=1, value=label)

for idx_r, row in enumerate(data):         # Iterate over the rows. Note idx_r starts at zero but openpyxl expect rows and columns starting at 1. Also account for labels above
    for idx_c, value in enumerate(row):
        if idx_c == 0 or idx_c==1 :
           # Write the 1st and 2nd column with date style
           sheet.cell(column=idx_c+1, row=idx_r+2, value=value)
	   cell=sheet.cell(column=idx_c+1, row=idx_r+2)  # So we can format the cells as date
	   cell.number_format = "mmm-dd-yy"
        else:
           sheet.cell(column=idx_c+1, row=idx_r+2, value=value)

# Adding a sheet for calculations
calc_sheet = bk.create_sheet(title="Calculations")
add_headers(calc_sheet,0)
# Adding sheet for google finance
google_sheet=bk.create_sheet(title="GoogleFinance")
add_headers(google_sheet,1)

# To obtain the current values
import json
from googlefinance import getQuotes

# Transaction types
trans_div_criteria="\"Dividend\""
trans_buy_criteria="\"Buy\""
trans_com_criteria="\"Buy\""
trans_tax_criteria="\"Tax\""

# write the list of the symbols

idx_row=2 # Start at row number 2
for value in symbols_list:

    # COL=1  Stock
    calc_sheet.cell(row=idx_row,column=1,value=value)
    google_sheet.cell(row=idx_row,column=1,value=value)
    #
    sym_criteria="A"+str(idx_row)

    # COL=2  Quantity Calculation (number of stocks held)
    calc_sheet.cell(row=idx_row,column=2,value="=SUMIF(%s,%s,%s)" %(symbols_range,sym_criteria,qty_range))
    google_sheet.cell(row=idx_row,column=2,value="=SUMIF(%s,%s,%s)" %(symbols_range,sym_criteria,qty_range))

    # COL=3  Dividends from Stock
    calc_sheet.cell(row=idx_row,column=3,value="=SUMIFS(%s,%s,%s,%s,%s)" %(amount_range,symbols_range,sym_criteria,trans_type_range,trans_div_criteria) )
    google_sheet.cell(row=idx_row,column=3,value="=SUMIFS(%s,%s,%s,%s,%s)" %(amount_range,symbols_range,sym_criteria,trans_type_range,trans_div_criteria) )

    dollar_format(calc_sheet.cell(row=idx_row,column=3))
    dollar_format(google_sheet.cell(row=idx_row,column=3))

    # COL=4  Amount of money Invested
    calc_sheet.cell(row=idx_row,column=4,value="=SUMIFS(%s,%s,%s,%s,%s)" %(amount_range,symbols_range,sym_criteria,trans_type_range,trans_buy_criteria) )
    google_sheet.cell(row=idx_row,column=4,value="=SUMIFS(%s,%s,%s,%s,%s)" %(amount_range,symbols_range,sym_criteria,trans_type_range,trans_buy_criteria) )
    dollar_format(calc_sheet.cell(row=idx_row,column=4))
    dollar_format(google_sheet.cell(row=idx_row,column=4))

    # COL=5 Commisions paid on buying/selling
    calc_sheet.cell(row=idx_row,column=5,value="=SUMIFS(%s,%s,%s,%s,%s)" %(com_range,symbols_range,sym_criteria,trans_type_range,trans_buy_criteria) )
    google_sheet.cell(row=idx_row,column=5,value="=SUMIFS(%s,%s,%s,%s,%s)" %(com_range,symbols_range,sym_criteria,trans_type_range,trans_buy_criteria) )

    dollar_format(calc_sheet.cell(row=idx_row,column=5))
    dollar_format(google_sheet.cell(row=idx_row,column=5))

    # COL=6  Taxes paid on the investment
    calc_sheet.cell(row=idx_row,column=6,value="=SUMIFS(%s,%s,%s,%s,%s)" %(amount_range,symbols_range,sym_criteria,trans_type_range,trans_tax_criteria) )
    dollar_format(calc_sheet.cell(row=idx_row,column=6))

    google_sheet.cell(row=idx_row,column=6,value="=SUMIFS(%s,%s,%s,%s,%s)" %(amount_range,symbols_range,sym_criteria,trans_type_range,trans_tax_criteria) )
    dollar_format(google_sheet.cell(row=idx_row,column=6))

    # COL=7  Calculation on dividend return vs. invested since stock was bought
    div_return_str="=C"+str(idx_row)+"/(-D"+str(idx_row)+"-F"+str(idx_row)+"+E"+str(idx_row)+")"
    calc_sheet.cell(row=idx_row,column=7,value=div_return_str )
    pct_format(calc_sheet.cell(row=idx_row,column=7))

    google_sheet.cell(row=idx_row,column=7,value=div_return_str )
    pct_format(google_sheet.cell(row=idx_row,column=7))

    # COL=8 Market. Very important for the price calculations
    # We will put the fully qualified symbol name here
    fq_symbol=dict_symbols[value]+":"+value
    calc_sheet.cell(row=idx_row,column=8,value=fq_symbol)
    google_sheet.cell(row=idx_row,column=8,value=fq_symbol)

    # COL=9 Current price obtained from the googlefinance API. For this need to determine which market the stock is on
    # obtain current price, should add TSE if CAD market
    if dict_symbols[value]=="TSE":
       fqsymbol="TSE:"+value
    else:              # Might need to check other stocks. For now assuming if not CAD everything else will be US market by default
       fqsymbol=value
    # This API broke. Alternative with my own but using opening price
    #quote=getQuotes([fqsymbol])[0] # getQuotes returns a list of dictionaries as one can pass multiple stocks at once
    #curr_price=quote['LastTradePrice'] # Will need to do some more later
    curr_price=getOpQuote(fqsymbol)
    curr_price=float(curr_price)
    calc_sheet.cell(row=idx_row,column=9,value=curr_price )
    dollar_format(calc_sheet.cell(row=idx_row,column=9))
    # For the google sheet we'll just insert the formula and let google populate the price
    market_str="H"+str(idx_row)
    google_sheet.cell(row=idx_row,column=9,value="=googlefinance(%s)" %(market_str))
    dollar_format(google_sheet.cell(row=idx_row,column=9))

    # COL=10 Price in CAD
    if dict_symbols[value]=="NYSEARCA":
       # Need to convert to CAD
       from forex_python.converter import CurrencyRates
       c = CurrencyRates()
       curr_price=c.convert("USD","CAD",curr_price)
    calc_sheet.cell(row=idx_row,column=10,value=curr_price )
    dollar_format(calc_sheet.cell(row=idx_row,column=10))
    # For google sheet we will use a formula. If currency of the stock is USD we should convert, otherwise leave untouched
    google_sheet_formula="=if(googlefinance(H"+str(idx_row)+",\"CURRENCY\")=\"USD\",googlefinance(\"CURRENCY:USDCAD\")*I"+str(idx_row)+",I"+str(idx_row)+")"
    google_sheet.cell(row=idx_row,column=10,value=google_sheet_formula)
    dollar_format(google_sheet.cell(row=idx_row,column=10))

    # COL=11 Current value of investmet. 
    amount_str="=B"+str(idx_row)+"*J"+str(idx_row)
    calc_sheet.cell(row=idx_row,column=11,value=amount_str)
    dollar_format(calc_sheet.cell(row=idx_row,column=11))
    google_sheet.cell(row=idx_row,column=11,value=amount_str)
    dollar_format(google_sheet.cell(row=idx_row,column=11))

    # COL=12 Estimated gain if the investment is sold. 
    # Gain is Current Price + Dividends obtained + Invested (it is negative) + Taxes (also negative) - Comision paid to buy - Comision to sell
    gain_str="=K"+str(idx_row)+"+C"+str(idx_row)+"+D"+str(idx_row)+"+F"+str(idx_row)+"-E"+str(idx_row)+COMISION
    calc_sheet.cell(row=idx_row,column=12,value=gain_str)
    google_sheet.cell(row=idx_row,column=12,value=gain_str)
    dollar_format(calc_sheet.cell(row=idx_row,column=12))
    dollar_format(google_sheet.cell(row=idx_row,column=12))

    # COL=13 Estimated return if the investment is sold. Calculate adding a comission (would need to be a constant)
    return_str="=L"+str(idx_row)+"/(-D"+str(idx_row)+"+E"+str(idx_row)+")"
    calc_sheet.cell(row=idx_row,column=13,value=return_str)
    google_sheet.cell(row=idx_row,column=13,value=return_str)
    pct_format(calc_sheet.cell(row=idx_row,column=13))
    pct_format(google_sheet.cell(row=idx_row,column=13))

    # COL=14 Name of the investment. Only doable on google finance
    name_str="=googlefinance(H"+str(idx_row)+",\"name\")"
    google_sheet.cell(row=idx_row,column=14,value=name_str)


    idx_row=idx_row+1

# Calling method to adjust the column widths
adjust_headers(calc_sheet)
adjust_headers(google_sheet)
# Add a procedure to calculate totals
totals_row=google_sheet.max_row
print "Max rows for google sheet"+ str(google_sheet.max_row)
write_totals(calc_sheet)
write_totals(google_sheet)


# Saving the created file
filesavename=os.path.splitext(filename)[0]+"_ord.xlsx"

print "Saving results to " + filesavename

bk.save(filesavename)# Notice this is xls, not xlsx like the original file is
